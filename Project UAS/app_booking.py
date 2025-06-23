import tkinter as tk
from tkinter import ttk, messagebox
import openpyxl
from openpyxl import Workbook
import os
from datetime import datetime

# --- Konfigurasi Awal ---
EXCEL_FILE = 'rekap_booking.xlsx'
DAYS = ['Senin', 'Selasa', 'Rabu', 'Kamis', 'Jumat', 'Sabtu']
ROOMS = ['B4A', 'B4B', 'B4H', 'B4C', 'B4E', 'B3B', 'B4R']
PRODI_LIST = ['Teknik Informatika', 'Sistem Informasi', 'Desain Komunikasi Visual', 'Manajemen']
SEMESTER_LIST = [str(i) for i in range(1, 9)] # Semester 1 sampai 8
GEDUNG_LIST = ['Gedung A', 'Gedung B', 'Gedung C']
LANTAI_LIST = ['Lantai 1', 'Lantai 2', 'Lantai 3', 'Lantai 4']
MODES = ['Offline', 'Online']


# --- Fungsi untuk Setup File Excel ---
# Fungsi ini akan membuat file Excel dengan header jika file belum ada
def setup_excel_file():
    if not os.path.exists(EXCEL_FILE):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Jadwal"
        # Menambahkan header baru sesuai permintaan
        headers = [
            "Nama Dosen", "Mata Kuliah", "SKS", "Kelas", "Prodi", "Semester", 
            "Hari", "Jam Mulai", "Jam Selesai", "Mode", 
            "Ruangan", "Gedung", "Lantai"
        ]
        sheet.append(headers)
        workbook.save(EXCEL_FILE)

# --- Kelas Utama Aplikasi ---
class BookingApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Aplikasi Booking Jadwal Kuliah")
        # Ukuran window diperbesar untuk menampung field baru
        self.root.geometry("500x750") 
        self.root.resizable(False, False)

        # Style
        self.style = ttk.Style(self.root)
        self.style.theme_use("clam")

        self.create_widgets()

    def create_widgets(self):
        # Frame utama
        main_frame = ttk.Frame(self.root, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # --- Frame Form Booking ---
        form_frame = ttk.LabelFrame(main_frame, text="Form Booking Jadwal Kuliah", padding="15")
        form_frame.pack(fill=tk.X, pady=(0, 20))
        form_frame.grid_columnconfigure(1, weight=1)

        # Nama Dosen
        ttk.Label(form_frame, text="Nama Dosen:").grid(row=0, column=0, sticky=tk.W, pady=5)
        self.nama_dosen_entry = ttk.Entry(form_frame)
        self.nama_dosen_entry.grid(row=0, column=1, sticky=tk.EW)

        # Mata Kuliah
        ttk.Label(form_frame, text="Mata Kuliah:").grid(row=1, column=0, sticky=tk.W, pady=5)
        self.matkul_entry = ttk.Entry(form_frame)
        self.matkul_entry.grid(row=1, column=1, sticky=tk.EW)

        # SKS
        ttk.Label(form_frame, text="SKS:").grid(row=2, column=0, sticky=tk.W, pady=5)
        self.sks_entry = ttk.Entry(form_frame)
        self.sks_entry.grid(row=2, column=1, sticky=tk.EW)

        # Kelas
        ttk.Label(form_frame, text="Kelas:").grid(row=3, column=0, sticky=tk.W, pady=5)
        self.kelas_entry = ttk.Entry(form_frame)
        self.kelas_entry.grid(row=3, column=1, sticky=tk.EW)
        
        # Prodi
        ttk.Label(form_frame, text="Program Studi:").grid(row=4, column=0, sticky=tk.W, pady=5)
        self.prodi_combobox = ttk.Combobox(form_frame, values=PRODI_LIST, state="readonly")
        self.prodi_combobox.grid(row=4, column=1, sticky=tk.EW)
        self.prodi_combobox.set('Pilih Prodi')

        # Semester
        ttk.Label(form_frame, text="Semester:").grid(row=5, column=0, sticky=tk.W, pady=5)
        self.semester_combobox = ttk.Combobox(form_frame, values=SEMESTER_LIST, state="readonly")
        self.semester_combobox.grid(row=5, column=1, sticky=tk.EW)
        self.semester_combobox.set('Pilih Semester')
        
        # Hari
        ttk.Label(form_frame, text="Hari:").grid(row=6, column=0, sticky=tk.W, pady=5)
        self.hari_combobox = ttk.Combobox(form_frame, values=DAYS, state="readonly")
        self.hari_combobox.grid(row=6, column=1, sticky=tk.EW)
        self.hari_combobox.set('Pilih Hari')

        # Jam Mulai
        ttk.Label(form_frame, text="Jam Mulai (HH:MM):").grid(row=7, column=0, sticky=tk.W, pady=5)
        self.jam_mulai_entry = ttk.Entry(form_frame)
        self.jam_mulai_entry.grid(row=7, column=1, sticky=tk.EW)

        # Jam Selesai
        ttk.Label(form_frame, text="Jam Selesai (HH:MM):").grid(row=8, column=0, sticky=tk.W, pady=5)
        self.jam_selesai_entry = ttk.Entry(form_frame)
        self.jam_selesai_entry.grid(row=8, column=1, sticky=tk.EW)
        
        # Mode
        ttk.Label(form_frame, text="Mode:").grid(row=9, column=0, sticky=tk.W, pady=5)
        self.mode_combobox = ttk.Combobox(form_frame, values=MODES, state="readonly")
        self.mode_combobox.grid(row=9, column=1, sticky=tk.EW)
        self.mode_combobox.set('Pilih Mode')
        # Bind event untuk menonaktifkan field lokasi jika mode 'Online'
        self.mode_combobox.bind("<<ComboboxSelected>>", self.toggle_location_fields)

        # Ruangan
        ttk.Label(form_frame, text="Ruangan:").grid(row=10, column=0, sticky=tk.W, pady=5)
        self.ruangan_combobox = ttk.Combobox(form_frame, values=ROOMS, state="readonly")
        self.ruangan_combobox.grid(row=10, column=1, sticky=tk.EW)
        self.ruangan_combobox.set('Pilih Ruangan')

        # Gedung
        ttk.Label(form_frame, text="Gedung:").grid(row=11, column=0, sticky=tk.W, pady=5)
        self.gedung_combobox = ttk.Combobox(form_frame, values=GEDUNG_LIST, state="readonly")
        self.gedung_combobox.grid(row=11, column=1, sticky=tk.EW)
        self.gedung_combobox.set('Pilih Gedung')
        
        # Lantai
        ttk.Label(form_frame, text="Lantai:").grid(row=12, column=0, sticky=tk.W, pady=5)
        self.lantai_combobox = ttk.Combobox(form_frame, values=LANTAI_LIST, state="readonly")
        self.lantai_combobox.grid(row=12, column=1, sticky=tk.EW)
        self.lantai_combobox.set('Pilih Lantai')
        
        # Tombol Simpan
        simpan_button = ttk.Button(form_frame, text="Simpan Booking", command=self.save_booking)
        simpan_button.grid(row=13, column=0, columnspan=2, pady=15)

        # --- Frame Lihat Jadwal ---
        view_frame = ttk.LabelFrame(main_frame, text="Lihat Jadwal Berdasarkan Ruangan", padding="15")
        view_frame.pack(fill=tk.X)
        view_frame.grid_columnconfigure(1, weight=1)

        ttk.Label(view_frame, text="Ruangan:").grid(row=0, column=0, sticky=tk.W, pady=5)
        self.view_ruangan_combobox = ttk.Combobox(view_frame, values=ROOMS, state="readonly")
        self.view_ruangan_combobox.grid(row=0, column=1, sticky=tk.EW, padx=(5,0))
        self.view_ruangan_combobox.set('Pilih Ruangan')

        tampilkan_button = ttk.Button(view_frame, text="Tampilkan Jadwal Ruangan", command=self.view_schedule)
        tampilkan_button.grid(row=1, column=0, columnspan=2, pady=10)

    def toggle_location_fields(self, event=None):
        """Menonaktifkan field lokasi jika mode 'Online' dipilih."""
        mode = self.mode_combobox.get()
        if mode == 'Online':
            # Kosongkan dan nonaktifkan
            self.ruangan_combobox.set('')
            self.gedung_combobox.set('')
            self.lantai_combobox.set('')
            self.ruangan_combobox.config(state=tk.DISABLED)
            self.gedung_combobox.config(state=tk.DISABLED)
            self.lantai_combobox.config(state=tk.DISABLED)
        else: # 'Offline'
            # Aktifkan kembali
            self.ruangan_combobox.config(state="readonly")
            self.gedung_combobox.config(state="readonly")
            self.lantai_combobox.config(state="readonly")
            self.ruangan_combobox.set('Pilih Ruangan')
            self.gedung_combobox.set('Pilih Gedung')
            self.lantai_combobox.set('Pilih Lantai')
            
    def clear_form(self):
        """Membersihkan semua field di form."""
        self.nama_dosen_entry.delete(0, tk.END)
        self.matkul_entry.delete(0, tk.END)
        self.sks_entry.delete(0, tk.END)
        self.kelas_entry.delete(0, tk.END)
        self.prodi_combobox.set('Pilih Prodi')
        self.semester_combobox.set('Pilih Semester')
        self.hari_combobox.set('Pilih Hari')
        self.jam_mulai_entry.delete(0, tk.END)
        self.jam_selesai_entry.delete(0, tk.END)
        self.mode_combobox.set('Pilih Mode')
        # Panggil toggle untuk memastikan status field lokasi benar
        self.toggle_location_fields()

    def check_conflict(self, hari, ruangan, jam_mulai_new, jam_selesai_new):
        """
        Pengecekan jadwal bentrok. Hanya berlaku untuk mode 'Offline'.
        Return True jika ada konflik, False jika tidak.
        """
        # Jika mode online (ruangan kosong), tidak akan ada konflik fisik
        if not ruangan or ruangan == 'Pilih Ruangan':
            return False

        workbook = openpyxl.load_workbook(EXCEL_FILE)
        sheet = workbook.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(row): continue
            
            # Indeks Kolom: 6=Hari, 7=Jam Mulai, 8=Jam Selesai, 10=Ruangan
            ex_hari, ex_jam_mulai, ex_jam_selesai, ex_ruangan = row[6], row[7], row[8], row[10]

            if ex_hari == hari and ex_ruangan == ruangan:
                try:
                    ex_start_time = datetime.strptime(str(ex_jam_mulai), "%H:%M").time()
                    ex_end_time = datetime.strptime(str(ex_jam_selesai), "%H:%M").time()
                    
                    if (jam_mulai_new < ex_end_time) and (jam_selesai_new > ex_start_time):
                        messagebox.showerror(
                            "Jadwal Bentrok!",
                            f"Ruangan {ruangan} sudah dibooking pada hari {hari} "
                            f"antara jam {ex_jam_mulai} - {ex_jam_selesai}."
                        )
                        return True # Ada konflik
                except (ValueError, TypeError):
                    continue
        
        return False # Tidak ada konflik

    def save_booking(self):
        """Mengambil data dari form dan menyimpannya ke Excel."""
        data = {
            "dosen": self.nama_dosen_entry.get(),
            "matkul": self.matkul_entry.get(),
            "sks": self.sks_entry.get(),
            "kelas": self.kelas_entry.get(),
            "prodi": self.prodi_combobox.get(),
            "semester": self.semester_combobox.get(),
            "hari": self.hari_combobox.get(),
            "mulai": self.jam_mulai_entry.get(),
            "selesai": self.jam_selesai_entry.get(),
            "mode": self.mode_combobox.get(),
            "ruangan": self.ruangan_combobox.get(),
            "gedung": self.gedung_combobox.get(),
            "lantai": self.lantai_combobox.get()
        }

        # Validasi Input Umum
        required_fields = ["dosen", "matkul", "sks", "kelas", "mulai", "selesai"]
        if any(not data[field] for field in required_fields) or \
           data['prodi'] == 'Pilih Prodi' or data['semester'] == 'Pilih Semester' or \
           data['hari'] == 'Pilih Hari' or data['mode'] == 'Pilih Mode':
            messagebox.showwarning("Input Tidak Lengkap", "Harap isi semua kolom yang wajib.")
            return

        # Validasi khusus untuk mode Offline
        if data['mode'] == 'Offline':
            if data['ruangan'] == 'Pilih Ruangan' or data['gedung'] == 'Pilih Gedung' or data['lantai'] == 'Pilih Lantai':
                messagebox.showwarning("Input Lokasi Tidak Lengkap", "Untuk mode Offline, harap isi Ruangan, Gedung, dan Lantai.")
                return
        else: # Mode Online
            # Set nilai lokasi ke N/A agar konsisten di Excel
            data['ruangan'] = 'N/A'
            data['gedung'] = 'N/A'
            data['lantai'] = 'N/A'
            
        try:
            jam_mulai_obj = datetime.strptime(data['mulai'], "%H:%M").time()
            jam_selesai_obj = datetime.strptime(data['selesai'], "%H:%M").time()
            if jam_mulai_obj >= jam_selesai_obj:
                messagebox.showerror("Waktu Tidak Valid", "Jam mulai harus sebelum jam selesai.")
                return
        except ValueError:
            messagebox.showerror("Format Waktu Salah", "Gunakan format HH:MM untuk jam (contoh: 14:00).")
            return

        # Cek jadwal bentrok (hanya jika mode Offline)
        if self.check_conflict(data['hari'], data['ruangan'], jam_mulai_obj, jam_selesai_obj):
            return

        # Simpan ke Excel
        try:
            workbook = openpyxl.load_workbook(EXCEL_FILE)
            sheet = workbook.active
            new_row = [
                data['dosen'], data['matkul'], data['sks'], data['kelas'], data['prodi'],
                data['semester'], data['hari'], data['mulai'], data['selesai'],
                data['mode'], data['ruangan'], data['gedung'], data['lantai']
            ]
            sheet.append(new_row)
            workbook.save(EXCEL_FILE)
            
            messagebox.showinfo("Sukses", "Booking jadwal berhasil disimpan!")
            self.clear_form()

        except Exception as e:
            messagebox.showerror("Error", f"Gagal menyimpan ke Excel: {e}")
            
    def view_schedule(self):
        """Menampilkan jadwal untuk ruangan yang dipilih di window baru."""
        selected_room = self.view_ruangan_combobox.get()
        if selected_room == 'Pilih Ruangan':
            messagebox.showwarning("Pilih Ruangan", "Silakan pilih ruangan untuk melihat jadwal.")
            return

        view_window = tk.Toplevel(self.root)
        view_window.title(f"Jadwal Ruangan {selected_room}")
        view_window.geometry("800x400") # Window lebih lebar untuk kolom tambahan

        columns = ("hari", "jam_mulai", "jam_selesai", "dosen", "matkul", "kelas", "prodi")
        tree = ttk.Treeview(view_window, columns=columns, show="headings")
        
        tree.heading("hari", text="Hari")
        tree.heading("jam_mulai", text="Jam Mulai")
        tree.heading("jam_selesai", text="Jam Selesai")
        tree.heading("dosen", text="Dosen")
        tree.heading("matkul", text="Mata Kuliah")
        tree.heading("kelas", text="Kelas")
        tree.heading("prodi", text="Prodi")
        
        tree.column("hari", width=80)
        tree.column("jam_mulai", width=80, anchor=tk.CENTER)
        tree.column("jam_selesai", width=80, anchor=tk.CENTER)
        tree.column("dosen", width=150)
        tree.column("matkul", width=150)
        tree.column("kelas", width=60, anchor=tk.CENTER)
        tree.column("prodi", width=150)
        
        try:
            workbook = openpyxl.load_workbook(EXCEL_FILE)
            sheet = workbook.active
            found_schedule = False
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not any(row): continue
                # Kolom Ruangan ada di indeks 10
                if row[10] == selected_room:
                    # Sesuaikan indeks untuk display_row
                    # row[6]=Hari, row[7]=Mulai, row[8]=Selesai, row[0]=Dosen, row[1]=Matkul, row[3]=Kelas, row[4]=Prodi
                    display_row = (row[6], row[7], row[8], row[0], row[1], row[3], row[4])
                    tree.insert("", tk.END, values=display_row)
                    found_schedule = True
            
            if not found_schedule:
                tree.insert("", tk.END, values=("", "", "Belum ada jadwal untuk ruangan ini.", "", "", "", ""))

        except Exception as e:
            messagebox.showerror("Error", f"Gagal membaca file Excel: {e}", parent=view_window)

        tree.pack(fill="both", expand=True, padx=10, pady=10)

# --- Blok Eksekusi Utama ---
if __name__ == "__main__":
    setup_excel_file()  # Pastikan file excel sudah siap
    root = tk.Tk()
    app = BookingApp(root)
    root.mainloop() 